from openpyxl import Workbook
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.drawing.image import Image
from report import *
import os
import sys
import pathlib
import glob  
import cv2
from ex import *
from ex1 import *

df=pd.read_csv("example_2.csv")
df=df.values.tolist()

def First():
	data=[]
	for i in range(0,len(df),5):
		temp=[]
		for j in range(1,11):
			temp+=[df[i][j]]
		data+=[temp]		    # data is a 2D list of individual candidate details
	sheet(data)                 # data is my candidate details
	

def sheet(data):		
	wb = openpyxl.load_workbook('Assignment.xlsx')
	ws = wb.active
	for i in range(0,len(data)):
	
		ws.title = str(i)
		ws['E7']=data[i][0]
		ws['I7']=data[i][1]
		ws['E8']=data[i][2]
		ws['I8']=data[i][3]
		ws['E9']=data[i][4]
		ws['I9']=data[i][5]
		ws['E10']=data[i][6]
		ws['I10']=data[i][7]
		ws['E11']=data[i][8]
		ws['I11']=data[i][9]
		spent_time,correct_score,your_score,incorrect_score,status_attempt,user_marked,answer_correct,coutcome,user_your_score=second(i)
		percentile,tscore=totalScore(your_score,i)
		ws['D22']=tscore
		ws['D24']=percentile
		ws['D16']=spent_time[0]   # Here spent_time is a 2D list where spent_time[0] will access the given 'i' students Time spent on Q1.
		ws['D17']=spent_time[1]   #Similarly, spent_time[1] will give time spent on Q2.
		ws['D18']=spent_time[2]   #Similarly is the case fo all variable below
		ws['D19']=spent_time[3]
		ws['D20']=spent_time[4]
		ws['E16']=correct_score[0]
		ws['E17']=correct_score[1]
		ws['E18']=correct_score[2]
		ws['E19']=correct_score[3]
		ws['E20']=correct_score[4]
		ws['F16']=incorrect_score[0]
		ws['F17']=incorrect_score[1]
		ws['F18']=incorrect_score[2]
		ws['F19']=incorrect_score[3]
		ws['F20']=incorrect_score[4]
		ws['G16']=status_attempt[0]
		ws['G17']=status_attempt[1]
		ws['G18']=status_attempt[2]
		ws['G19']=status_attempt[3]
		ws['G20']=status_attempt[4]
		ws['H16']=user_marked[0]
		ws['H17']=user_marked[1]
		ws['H18']=user_marked[2]
		ws['H19']=user_marked[3]
		ws['H20']=user_marked[4]
		ws['I16']=answer_correct[0]
		ws['I17']=answer_correct[1]
		ws['I18']=answer_correct[2]
		ws['I19']=answer_correct[3]
		ws['I20']=answer_correct[4]
		ws['J16']=coutcome[0]
		ws['J17']=coutcome[1]
		ws['J18']=coutcome[2]
		ws['J19']=coutcome[3]
		ws['J20']=coutcome[4]
		ws['K16']=user_your_score[0]
		ws['K17']=user_your_score[1]
		ws['K18']=user_your_score[2]
		ws['K19']=user_your_score[3]
		ws['K20']=user_your_score[4]
		wb.save("outfile"+str(i)+'.xlsx')
	final_files()
	 	
		
if __name__ == "__main__" :
	First()
		
		
		